#! python3
# sendDuesReminders.py - スプレッドシートの支払い状況に基づきメールを送信

import openpyxl, smtplib, sys

# スプレッドシートを開き最近の支払い状況を取得
wb = openpyxl.load_workbook('duesRecords.xlsx')  # ❶
sheet = wb.get_sheet_by_name('Sheet1')  # ❷

last_col = sheet.get_highest_column()  # ❸
latest_month = sheet.cell(row=1, column=last_col).value  # ❹

# 会員の支払い状況を調べる
unpaid_members = {}
for r in range(2, sheet.get_highest_row() + 1):       # ❶
    payment = sheet.cell(row=r, column=last_col).value # ❷
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value      # ❸
        email = sheet.cell(row=r, column=2).value     # ❹
        unpaid_members[name] = email  # ❺

# メールアカウントにログインする
smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)
smtp_obj.ehlo()
smtp_obj.starttls()
smtp_obj.login('my_email_address@gmail.com', sys.argv[1])

# TODO: リマインダーメールを送信する
for name, email in unpaidMembers.items():
    body = """Subject: {} dues unpaid.
Dear {},
Records show that you have not paid dues for {}. Please make this payment as soon as possible. Thank you!
""".format(latest_month, name, latest_month)  # ❶
    print('メール送信中 {}...'.format(email))  # ❷
    sendmail_status = smtp_obj.sendmail('my_email_address@gmail.com', email, body)  # ❸

    if sendmail_status != {}:  # ❹
        print('{}へメール送信中に問題が起こりました: {}'.format(email, endmail_status))

smtp_obj.quit()

